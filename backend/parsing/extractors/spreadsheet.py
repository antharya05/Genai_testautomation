"""Excel (.xlsx) and CSV extractors.

Both formats are inherently tabular, so each worksheet / file becomes one
normalized table that the structured stage column-detects. CSV uses the stdlib
``csv`` module with delimiter sniffing; XLSX uses openpyxl (already a project
dependency).
"""

from __future__ import annotations

import csv

from . import ExtractedContent, Table


def extract_xlsx(path: str) -> ExtractedContent:
    from openpyxl import load_workbook

    content = ExtractedContent()
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        for ws in wb.worksheets:
            rows: Table = []
            for row in ws.iter_rows(values_only=True):
                cells = ["" if v is None else str(v).strip() for v in row]
                if any(cells):
                    rows.append(cells)
            if rows:
                content.tables.append(rows)
    finally:
        wb.close()

    if not content.tables:
        content.issues.append("No non-empty worksheets found in workbook.")
    return content


def extract_csv(path: str) -> ExtractedContent:
    content = ExtractedContent()
    with open(path, "r", encoding="utf-8-sig", errors="replace", newline="") as fh:
        sample = fh.read(4096)
        fh.seek(0)
        delimiter = _sniff_delimiter(sample)
        reader = csv.reader(fh, delimiter=delimiter)
        rows: Table = [
            [c.strip() for c in row]
            for row in reader
            if any(c.strip() for c in row)
        ]
    if rows:
        content.tables.append(rows)
    else:
        content.issues.append("CSV file contained no data rows.")
    return content


def _sniff_delimiter(sample: str) -> str:
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        return dialect.delimiter
    except csv.Error:
        return ","
