import csv
import io
from datetime import datetime


def _asil_hex(asil: str) -> str:
    return {
        "QM": "F1F5F9", "A": "D1FAE5",
        "B": "FEF9C3", "C": "FFEDD5", "D": "FEE2E2",
    }.get(asil, "FFFFFF")


def _manifest_rows(manifest: dict) -> list[tuple[str, str]]:
    """Flatten an approval manifest into (label, value) rows for embedding."""
    s = manifest.get("summary", {}) or {}
    return [
        ("Run ID", str(manifest.get("run_id", ""))),
        ("Project", str(manifest.get("project_name", ""))),
        ("Review State", str(manifest.get("review_state", ""))),
        ("Locked", "yes" if manifest.get("locked") else "no"),
        ("Approved By", str(manifest.get("approved_by_display") or "")),
        ("Approved At", str(manifest.get("approved_at") or "")),
        ("Approved / Total", f"{s.get('approved', 0)} / {s.get('total', 0)}"),
        ("Approved %", str(s.get("approved_pct", 0))),
        ("Content Digest", str(manifest.get("review_digest") or "")),
        ("Stale (content changed since sign-off)", "yes" if manifest.get("stale") else "no"),
        ("Exported At", datetime.utcnow().isoformat()),
    ]


def export_excel(
    test_cases: list[dict],
    project_name: str = "automotive_project",
    manifest: dict | None = None,
) -> bytes:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError("openpyxl not installed. Run: pip install openpyxl") from exc

    wb = openpyxl.Workbook()

    # ── Sheet 1: Test Cases ──────────────────────────────────────
    ws = wb.active
    ws.title = "Test Cases"

    hdr_font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
    hdr_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    top_left = Alignment(horizontal="left", vertical="top", wrap_text=True)
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = [
        "Test ID", "Requirement ID", "Title", "ASIL", "ASIL Source", "ASIL Confidence",
        "Test Type", "Review Status", "Reviewer", "Reviewed At", "Preconditions", "Steps",
        "Expected Results", "Model", "Prompt Ver", "Timestamp", "Retries",
    ]
    widths = [12, 18, 42, 8, 13, 14, 16, 15, 18, 22, 40, 52, 52, 22, 12, 26, 8]

    for col, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = center
        c.border = thin
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    for ridx, tc in enumerate(test_cases, 2):
        asil = tc.get("asil", "QM")
        row_fill = PatternFill(
            start_color=_asil_hex(asil),
            end_color=_asil_hex(asil),
            fill_type="solid",
        )
        steps_text = "\n".join(
            f"{i + 1}. {s}" for i, s in enumerate(tc.get("steps", []))
        )
        values = [
            tc.get("test_id", ""),
            tc.get("requirement_id", ""),
            tc.get("title", ""),
            asil,
            tc.get("asil_source", "estimated"),
            str(tc.get("asil_confidence", 100)),
            tc.get("test_type", ""),
            tc.get("review_status", "pending"),
            tc.get("reviewer", ""),
            tc.get("reviewed_at", ""),
            "\n".join(f"• {p}" for p in tc.get("preconditions", [])),
            steps_text,
            "\n".join(f"• {r}" for r in tc.get("expected_results", [])),
            tc.get("model_version", ""),
            tc.get("prompt_version", "v1"),
            tc.get("generation_timestamp", ""),
            str(tc.get("retry_count", 0)),
        ]
        for col, val in enumerate(values, 1):
            c = ws.cell(row=ridx, column=col, value=val)
            c.alignment = top_left if col > 7 else center
            c.border = thin
            c.fill = row_fill

        step_lines = len(steps_text.splitlines()) if steps_text else 1
        ws.row_dimensions[ridx].height = max(30, step_lines * 15)

    # ── Sheet 2: Traceability Matrix ─────────────────────────────
    ws2 = wb.create_sheet("Traceability Matrix")
    bold = Font(bold=True, name="Calibri")
    for col, h in enumerate(["Requirement ID", "Linked Test Cases", "Coverage Count"], 1):
        ws2.cell(row=1, column=col, value=h).font = bold
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 55
    ws2.column_dimensions["C"].width = 18

    req_map: dict[str, list[str]] = {}
    for tc in test_cases:
        req_map.setdefault(tc.get("requirement_id", "REQ_UNKNOWN"), []).append(
            tc.get("test_id", "")
        )

    for ridx, (req_id, tc_ids) in enumerate(sorted(req_map.items()), 2):
        ws2.cell(row=ridx, column=1, value=req_id)
        ws2.cell(row=ridx, column=2, value=", ".join(tc_ids))
        ws2.cell(row=ridx, column=3, value=len(tc_ids))

    # ── Sheet 3: Approval Manifest (governance provenance) ───────
    if manifest:
        ws3 = wb.create_sheet("Approval Manifest")
        ws3.column_dimensions["A"].width = 38
        ws3.column_dimensions["B"].width = 60
        ws3.cell(row=1, column=1, value="Field").font = bold
        ws3.cell(row=1, column=2, value="Value").font = bold
        for ridx, (label, val) in enumerate(_manifest_rows(manifest), 2):
            ws3.cell(row=ridx, column=1, value=label).font = bold
            ws3.cell(row=ridx, column=2, value=val)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def export_jira_csv(test_cases: list[dict], manifest: dict | None = None) -> str:
    _priority = {"QM": "Low", "A": "Medium", "B": "Medium", "C": "High", "D": "Critical"}

    out = io.StringIO()
    # Provenance header: comment lines (a common, importer-tolerant convention) so
    # an approved CSV artifact is self-describing about who signed it off.
    if manifest:
        for label, val in _manifest_rows(manifest):
            out.write(f"# {label}: {val}\n")
    fieldnames = [
        "Issue Type", "Summary", "Description", "Priority",
        "Labels", "Custom field (Test Type)",
        "Custom field (ASIL Level)", "Custom field (ASIL Source)",
        "Custom field (ASIL Confidence)", "Custom field (Requirement ID)",
        "Custom field (Review Status)", "Custom field (Reviewer)",
        "Custom field (Reviewed At)",
    ]
    writer = csv.DictWriter(out, fieldnames=fieldnames)
    writer.writeheader()

    for tc in test_cases:
        asil = tc.get("asil", "QM")
        precond = "\n".join(f"• {p}" for p in tc.get("preconditions", []))
        steps = "\n".join(
            f"{i + 1}. {s}" for i, s in enumerate(tc.get("steps", []))
        )
        results = "\n".join(f"• {r}" for r in tc.get("expected_results", []))
        description = (
            f"*Preconditions:*\n{precond}\n\n"
            f"*Test Steps:*\n{steps}\n\n"
            f"*Expected Results:*\n{results}"
        )
        writer.writerow({
            "Issue Type": "Test",
            "Summary": f"{tc.get('test_id', '')} - {tc.get('title', '')}",
            "Description": description,
            "Priority": _priority.get(asil, "Medium"),
            "Labels": f"automotive,iso26262,{asil.lower()}",
            "Custom field (Test Type)": tc.get("test_type", ""),
            "Custom field (ASIL Level)": asil,
            "Custom field (ASIL Source)": tc.get("asil_source", "estimated"),
            "Custom field (ASIL Confidence)": tc.get("asil_confidence", 100),
            "Custom field (Requirement ID)": tc.get("requirement_id", ""),
            "Custom field (Review Status)": tc.get("review_status", "pending"),
            "Custom field (Reviewer)": tc.get("reviewer", ""),
            "Custom field (Reviewed At)": tc.get("reviewed_at", ""),
        })

    return out.getvalue()
